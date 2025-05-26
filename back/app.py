from flask import Flask, send_file, render_template, request, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import os
#from ./class/class import Excel

app = Flask(__name__)

#@app.route("/")
#def hello_world():
#    return "<p>Hello, World!</p>"

@app.get("/")
def upload():
    return render_template('upload.html')

@app.post('/test')
def test():
    file1 = request.files['one'] # 'one' es el nombre del input en el formulario
    file2 = request.files['two']
    data_file1 = openpyxl.load_workbook(file1)
    data_file2 = openpyxl.load_workbook(file2)

    data_sheet1 = data_file1["Sheet1"] # Nombre de la hoja que quieres leer
    data_sheet2 = data_file2["Sheet1"]

    for row in data_sheet1.iter_rows():
        for cell in row:
            current_cell_value = cell.value
            cell_location = cell.coordinate

            if current_cell_value != data_sheet2[cell_location].value:
                print(data_sheet2[cell_location])
    try:
        df1 = pd.read_excel(data_file1, engine='openpyxl')
        df2 = pd.read_excel(data_file2, engine='openpyxl')
        excel_file = pd.read_excel(file1, engine='openpyxl')
    except Exception as e:
        return f"Error al procesar los archivos: {str(e)}", 400


    return f"""
    <html>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.6/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-4Q6Gf2aSP4eDXB8Miphtr37CMZZQ5oXLH2yaXMJ2w8e2ZtHTl7GptT4jmndRuHDT" crossorigin="anonymous">
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.6/dist/js/bootstrap.bundle.min.js" integrity="sha384-j1CDi7MgGQ12Z7Qab0qlWQ/Qqz24Gc6BM0thvEMVjHnfYGF0rmFCozFSxQBxwHKO" crossorigin="anonymous"></script>
    <head><title>Vista de archivos Excel</title></head>
    <body>
    <div class="row">
    <div class="col-md-5">
    <h2>Contenido del archivo 1: {file1.filename}</h2>
    {df1.to_html(classes="table table-striped table-hover table-bordered")}
    </div>
    <div class="col-md-5">
    <h2>Contenido del archivo 2: {file2.filename}</h2>
    {df2.to_html(classes="table table-striped table-hover table-bordered")}
    </div>
    </div>
    </body>
    </html>
    """





@app.post('/view')
def view():
    newExcel = Excel
    file1 = request.files['one'] # 'one' es el nombre del input en el formulario
    file2 = request.files['two']

    if not file1.filename.endswith('.xlsx') or not file2.filename.endswith('.xlsx'):
        return "Ambos archivos deben ser .xlsx", 400

    try:
        df1 = pd.read_excel(file1, engine='openpyxl')
        df2 = pd.read_excel(file2, engine='openpyxl')
        excel_file = pd.read_excel(file1, engine='openpyxl')
    except Exception as e:
        return f"Error al procesar los archivos: {str(e)}", 400

    return f"""
    <html>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.6/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-4Q6Gf2aSP4eDXB8Miphtr37CMZZQ5oXLH2yaXMJ2w8e2ZtHTl7GptT4jmndRuHDT" crossorigin="anonymous">
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.6/dist/js/bootstrap.bundle.min.js" integrity="sha384-j1CDi7MgGQ12Z7Qab0qlWQ/Qqz24Gc6BM0thvEMVjHnfYGF0rmFCozFSxQBxwHKO" crossorigin="anonymous"></script>
    <head><title>Vista de archivos Excel</title></head>
    <body>
    <div class="row">
    <div class="col-md-5">
    <h2>Contenido del archivo 1: {file1.filename}</h2>
    {df1.to_html(classes="table table-bordered")}
    </div>
    <div class="col-md-5">
    <h2>Contenido del archivo 2: {file2.filename}</h2>
    {df2.to_html(classes="table table-bordered")}
    </div>
    </div>
    </body>
    </html>
    """, print(excel_file.columns), print(excel_file.dtypes) #print(excel_file.head())  #print(excel_file.info()) #print(excel_file.describe())



@app.route("/download")
def health_check():
    # Cargar el DataFrame que quieres insertar
    tips = pd.read_csv("https://raw.githubusercontent.com/mwaskom/seaborn-data/master/tips.csv")

    # Cargar plantilla Excel
    plantilla_path = "./example/Test.xlsx"
    wb = load_workbook(plantilla_path)
    ws = wb.active  # Puedes elegir otra hoja con wb["SheetName"]

    # Escribir los datos en la hoja (a partir de la fila 2, por ejemplo)
    for i, row in tips.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)  # Ajusta según el formato de tu plantilla
        if i >= 10:  # Solo agregar las primeras 10 filas para no llenar demasiado
            break

    # Guardar a un archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     as_attachment=True,
                     download_name="reporte_con_datos.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int("8080"), debug=True)